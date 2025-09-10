import re
import pandas as pd
from uniqueid import get_uid_length  # 导入 uniqueid 模块中的函数
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import shutil
from openpyxl.comments import Comment
from openpyxl.styles import Font

def parse_otp_info(otp_str, file_content, flash_name):
    """解析OTP信息并返回描述"""
    if not otp_str or 'NULL' in otp_str:
        return ''
        
    # 只为 EN25QE32A 打印调试信息
    debug_print = (flash_name == 'EN25QE32A')
    
    # 从otp_str中提取类型编号（例如从otp_type18中提取18）
    otp_type_match = re.search(r'otp_type(\d+)', otp_str)
    if not otp_type_match:
        return ''
        
    otp_type_num = otp_type_match.group(1)
    if debug_print:
        print(f"\nEN25QE32A OTP type number: {otp_type_num}")
    
    # 在文件内容中查找对应的OTP定义
    pattern = rf'static\s+struct\s+spi_nor_otp\s+otp_type{otp_type_num}\s*=\s*\{{([^}}]+)\}}'
    otp_def_match = re.search(pattern, file_content)
    if not otp_def_match:
        if debug_print:
            print(f"Could not find OTP definition for type {otp_type_num}")
        return ''
    
    if debug_print:
        print(f"EN25QE32A OTP definition: {otp_def_match.group(0)}")
    
    # 解析OTP定义中的值
    otp_values = otp_def_match.group(1).split(',')
    if len(otp_values) >= 4:
        try:
            # 注意：最后两个值的顺序是 regions, size
            regions = int(otp_values[-2].strip())  # 倒数第二个值是 regions
            size = int(otp_values[-3].strip())     # 倒数第三个值是 size
            if debug_print:
                print(f"EN25QE32A OTP parsed values: size={size}, regions={regions}")
            return f"{regions}*{size} Byte"  # 修改格式为 regions*sizeBytes
        except ValueError as e:
            if debug_print:
                print(f"Error parsing OTP values: {e}")
            return ''
            
    return ''

def parse_flash_info(file_content):
    # 用于存储所有flash信息的列表
    flash_list = []
    
    # 使用正则表达式匹配spi_nor_info结构体数组中的每个元素
    pattern = r'\{\s*"([^"]+)",\s*' # Flash名称
    pattern += r'0x([0-9a-f]+),\s*' # JEDEC ID
    pattern += r'0x([0-9a-f]+),\s*' # 容量
    pattern += r'([^,]+),\s*' # 特性
    pattern += r'([^,]+),\s*' # get_flash_markid
    pattern += r'([^,]+),\s*' # markid
    pattern += r'([^,]+),\s*' # protect
    pattern += r'([^}]+)\s*\},' # otp
    
    matches = re.finditer(pattern, file_content, re.MULTILINE | re.DOTALL)
    
    for match in matches:
        flash_name = match.group(1)
        jedec_id_str = match.group(2)
        size = match.group(3)
        features = match.group(4).strip()
        get_markid = match.group(5).strip()
        markid = match.group(6).strip()
        protect = match.group(7).strip()
        otp = match.group(8).strip() if match.group(8) else ''
        
        # 将jedec_id转换为整数
        jedec_id = int(jedec_id_str, 16)
        
        # 使用 uniqueid 模块中的函数计算UID长度
        uid_len = get_uid_length(jedec_id, markid)
        
        # 解析OTP信息，传入flash_name
        otp_desc = parse_otp_info(otp, file_content, flash_name)
        
        # 转换容量为MB
        size_mb = int(size, 16) / (1024 * 1024)
        
        # 解析特性
        feature_list = []
        if 'SPI_NOR_DUAL_READ' in features:
            feature_list.append('DUAL_READ')
        if 'SPI_NOR_QUAD_READ' in features:
            feature_list.append('QUAD_READ')
        if 'SPI_NOR_QUAD_WRITE' in features:
            feature_list.append('QUAD_WRITE')
        features_str = '/'.join(feature_list) if feature_list else 'None'
        
        # 解析protect类型
        protect_type = 'None'
        if protect and 'protect_' in protect and protect != 'NULL':  # 添加 protect != 'NULL' 的检查
            protect_match = re.search(r'&protect_([a-z0-9_]+)', protect)
            if protect_match:
                protect_type = protect_match.group(1)
                
        # 解析OTP类型
        otp_type = 'None'
        if 'otp_type' in otp:
            otp_match = re.search(r'otp_type(\d+)', otp)
            if otp_match:
                otp_type = f'type{otp_match.group(1)}'
        
        flash_list.append({
            'Flash Name': flash_name,
            'JEDEC ID': f'0x{jedec_id_str}',
            'Size(MB)': size_mb,
            'Features': features_str,
            'Get MarkID': 'Yes' if get_markid != 'NULL' else 'No',
            'Mark ID': markid if markid != '0' else 'None',
            'Protect Type': protect_type,  # 这里存储的应该是完整的类型名称
            'OTP': otp_desc
        })
    
    return flash_list

def get_brand_name(flash_name, jedec_id):
    """根据Flash名称和JEDEC ID确定品牌"""
    # 特殊情况处理
    if flash_name == 'FM25Q64A' and jedec_id.lower() == '0x00f83217':
        return 'Doslicon(东芯)'
    
    # 获取前缀（1-2个字符）
    prefix = flash_name[:2] if len(flash_name) >= 2 else flash_name[:1]
    
    # 品牌映射字典
    brand_map = {
        'EN': 'Eon(宜杨)',
        'F': 'ESMT(晶豪科技)',
        'GD': 'GigaDevice(兆易创新)',
        'M': 'Micron(镁光)',
        'N': 'Micron(镁光)',
        'MX': 'MXIC(旺宏)',
        'W': 'Winbond(华邦)',
        'S': 'Cypress(赛普拉斯)',
        'PM': 'PMC(常亿)',
        'PN': 'Paragon(芯天下曾用名)',
        'IC': 'ISSI(芯成)',
        'ZB': 'Zbit(恒烁)',
        'BY': 'Boya(博雅)',
        'BH': 'Boya(博雅)',
        'XM': 'XMC(武汉新芯)',
        'FM': 'FUDAN(复旦微)',
        'FS': 'FORESEE(江波龙)',
        'XT': 'XTX(芯天下)',
        'ZD': 'Zetta(芯泽)',
        'NM': 'NORMEM(诺存)',
        'DS': 'Doslicon(东芯)',
        'JY': 'Juyang(巨扬)',
        'UC': 'Ucun(优存科技)',
        'PY': 'Puya(普冉)',
        'T': 'First-Rank(飞思瑞克)',
        'SX': 'SANXIA(三峡)',
        'GT': 'GT(聚辰)',
        'TH': 'TH(紫光青藤)',
        'BW': 'BW(佰维)'
    }
    
    # 先尝试两个字符的前缀
    if prefix in brand_map:
        return brand_map[prefix]
    
    # 如果两个字符没找到，且前缀是两个字符，尝试第一个字符
    if len(prefix) == 2 and prefix[0] in brand_map:
        return brand_map[prefix[0]]
    
    # 如果都没找到，返回原始名称的第一个单词
    return flash_name.split()[0]

def analyze_excel_structure(excel_file):
    """解析Excel表格结构"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 获取表头信息
        header_row = 1
        subheader_row = 2
        columns = {}
        
        print("\n=== Excel表格结构分析 ===")
        print("列号\t主标题\t子标题")
        print("-" * 40)
        
        for col in range(1, ws.max_column + 1):
            main_header = ws.cell(row=header_row, column=col).value
            sub_header = ws.cell(row=subheader_row, column=col).value
            
            if main_header or sub_header:
                columns[col] = {
                    'main_header': main_header,
                    'sub_header': sub_header
                }
                print(f"{col}\t{main_header}\t{sub_header}")
        
        return columns
        
    except Exception as e:
        print(f"解析Excel结构时发生错误: {str(e)}")
        return None

def prepare_new_flash_info(flash_data, excel_columns):
    """根据Excel格式准备新的flash信息"""
    # 获取品牌信息
    brand = get_brand_name(flash_data['Flash Name'], flash_data['JEDEC ID'])
    
    # 解析JEDEC ID获取MID和DID
    jedec_id = flash_data['JEDEC ID'].lower()
    if len(jedec_id) >= 10:
        jedec_id = "0x" + jedec_id[4:]
    mid = f"0x{jedec_id[2:4].upper()}"
    did = f"0x{jedec_id[4:8].upper()}"
    
    # 解析速度支持
    features = str(flash_data['Features'])
    if ('DUAL_READ' in features and 
        'QUAD_READ' in features and 
        'QUAD_WRITE' in features):
        speed_support_str = '支持'
    elif 'DUAL_READ' in features:
        speed_support_str = '不支持4倍速'
    else:
        speed_support_str = '不支持2、4倍速'
    
    # 准备完整的flash信息
    formatted_info = {}
    for col, headers in excel_columns.items():
        main_header = headers['main_header']
        sub_header = headers['sub_header']
        
        # 如果主标题和子标题都为空，跳过该列
        if not main_header and not sub_header:
            continue
            
        # 根据列标题填充对应的值
        value = ''
        if main_header == '品牌':
            value = brand
        elif main_header == '主型号':
            value = flash_data['Flash Name']
        elif main_header == 'MID':
            value = mid
        elif main_header == 'DID':
            value = did
        elif main_header == '容量\n(MB)':
            value = int(flash_data['Size(MB)'])
        elif main_header == '是否支持多倍速':
            value = speed_support_str
        elif main_header == 'OTP':
            value = flash_data['OTP']
        elif main_header == '读写擦':
            value = '支持'
        elif main_header == '写保护':
            value = '支持' if flash_data.get('Protect Type') and flash_data.get('Protect Type') != 'None' else '不支持'
        elif main_header in ['料号', '封装', '最高时钟频率', '备注', 'UID']:
            value = ''  # 这些列保持空白
        else:
            # 对于其他非空列（有效列），填入"不支持"
            value = '不支持'
        
        formatted_info[col] = value
    
    return formatted_info

def parse_protect_range(protect_str, file_content):
    """解析写保护范围并返回描述"""
    if not protect_str or protect_str == 'None' or protect_str == 'NULL':
        return ''
    
    # 在文件内容中查找对应的protect range数组定义
    pattern = rf'static\s+struct\s+spi_nor_flash_protect_range\s+protect_range_{protect_str}\[\]\s*=\s*\{{([^;]+)\}};'
    protect_def_match = re.search(pattern, file_content, re.DOTALL)
    if not protect_def_match:
        return ''
    
    array_content = protect_def_match.group(1)
    
    # 解析每��护范围条目
    ranges = []
    
    # 使用正则表达式匹配每个条目中的块大小
    entries = re.finditer(r'\{\s*0x[0-9a-fA-F]+,\s*0x[0-9a-fA-F]+,\s*0x[0-9a-fA-F]+,\s*0x[0-9a-fA-F]+,\s*(\d+)\s*\*\s*0x10000\s*\}', array_content)
    
    for entry in entries:
        try:
            blocks = int(entry.group(1))
            if blocks == 0:
                end_addr = 0
            else:
                end_addr = blocks * 0x10000 - 1
            
            range_str = f"{blocks:2d} blocks   : 0x00 ~ 0x{end_addr:06X}"
            ranges.append(range_str)
        except (ValueError, IndexError):
            continue
    
    # 按块数排序
    ranges.sort(key=lambda x: int(x.split()[0]))
    
    return '\n'.join(ranges)

def compare_and_update_excel(flash_list, existing_excel, output_excel):
    try:
        print("\n=== 开始处理Excel文件 ===")
        
        # 读取源文件内容
        print("正在读取源文件...")
        with open('drivers/spinor/spinor_ids.c', 'r', encoding='utf-8') as f:
            source_content = f.read()
        print("源文件读取成功")
        
        import shutil
        shutil.copy2(existing_excel, output_excel)
        
        wb = openpyxl.load_workbook(output_excel)
        ws = wb.active
        
        # 定义头��
        header_row = 1
        subheader_row = 2
        data_start_row = 3
        
        # 找到第一个OTP列、写保护列和读写擦列
        first_otp_column = None
        first_wp_column = None
        first_rwe_column = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=subheader_row, column=col).value
            main_header = ws.cell(row=header_row, column=col).value
            
            if cell_value:
                cell_value = str(cell_value).strip()
                if cell_value == 'OTP' and not first_otp_column:
                    first_otp_column = col
                    print(f"Found first OTP column: {col}")
                elif cell_value == '写保护' and not first_wp_column:
                    first_wp_column = col
                    print(f"Found first Write Protect column: {col}")
                elif cell_value == '读写擦' and not first_rwe_column:
                    first_rwe_column = col
                    print(f"Found first Read/Write/Erase column: {col}")
        
        # 打印找到的列
        print(f"\nFound columns:")
        print(f"OTP column: {first_otp_column}")
        print(f"Write Protect column: {first_wp_column}")
        print(f"Read/Write/Erase column: {first_rwe_column}")
        
        if not first_otp_column or not first_wp_column or not first_rwe_column:
            print("Warning: Required columns not found!")
            return
            
        # 获取现有的MID-DID组合
        existing_mid_did = set()
        for row in range(data_start_row, ws.max_row + 1):
            mid = str(ws.cell(row=row, column=5).value).lower()
            did = str(ws.cell(row=row, column=6).value).lower()
            if mid != "none" and did != "none":
                existing_mid_did.add(f"{mid}-{did}")
        
        # 处理新的Flash设备
        new_rows = []
        for flash in flash_list:
            jedec_id = flash['JEDEC ID'].lower()
            if len(jedec_id) >= 10:
                jedec_id = "0x" + jedec_id[4:]
            
            mid = f"0x{jedec_id[2:4].upper()}"
            did = f"0x{jedec_id[4:8].upper()}"
            
            mid_did = f"{mid.lower()}-{did.lower()}"
            if mid_did in existing_mid_did:
                continue
            
            brand = get_brand_name(flash['Flash Name'], flash['JEDEC ID'])
            
            # 定义speed_support_str
            features = str(flash['Features'])
            if ('DUAL_READ' in features and 
                'QUAD_READ' in features and 
                'QUAD_WRITE' in features):
                speed_support_str = '支持'
            elif 'DUAL_READ' in features:
                speed_support_str = '不支持4倍速'
            else:
                speed_support_str = '不支持2、4倍速'
            
            # 保存写保护类型信息
            protect_type = flash.get('Protect Type', '')
            
            new_rows.append({
                '品牌': brand,
                '主型号': flash['Flash Name'],
                '料号': '',
                '封装': '',
                'MID': mid,
                'DID': did,
                '容量\n(MB)': int(flash['Size(MB)']),
                '是否支持多倍速': speed_support_str,
                '最高时钟频率': '',
                '备注': '',
                'OTP': flash['OTP'] if 'OTP' in flash else '',
                'protect_type': protect_type,  # 添加写保护类型
                'protect_ranges': parse_protect_range(protect_type, source_content) if protect_type else ''  # 预先解析写保护范围
            })
        
        if not new_rows:
            print("没有发现新的Flash设备")
            return
        
        # 按品牌分组
        brand_groups = {}
        for new_row in new_rows:
            brand = new_row['品牌']
            if brand not in brand_groups:
                brand_groups[brand] = []
            brand_groups[brand].append(new_row)
        
        # 在相应品牌后插入新行
        current_row = data_start_row
        for row in range(data_start_row, ws.max_row + 1):
            brand = ws.cell(row=row, column=1).value
            if brand in brand_groups:
                # 找到品牌的最后一行
                last_row = row
                while (last_row < ws.max_row and 
                       ws.cell(row=last_row + 1, column=1).value == brand):
                    last_row += 1
                
                # 在品牌的最后插入新行
                for new_row in brand_groups[brand]:
                    current_row = last_row + 1
                    ws.insert_rows(current_row)
                    
                    # 复制上一行的样式和设置值
                    for col in range(1, ws.max_column + 1):
                        new_cell = ws.cell(row=current_row, column=col)
                        old_cell = ws.cell(row=last_row, column=col)
                        
                        # 获取列标题
                        main_header = ws.cell(row=header_row, column=col).value
                        sub_header = ws.cell(row=subheader_row, column=col).value
                        
                        # 复制样式
                        if old_cell.has_style:
                            new_cell._style = copy(old_cell._style)
                        
                        # 如果主标题和子标题都为空，跳过该列
                        if not main_header and not sub_header:
                            continue
                        
                        # 设置值
                        if col == first_otp_column:
                            if new_row['OTP']:
                                new_cell.value = new_row['OTP']
                        elif col == first_wp_column:
                            protect_type = new_row.get('protect_type', '')
                            protect_ranges = new_row.get('protect_ranges', '')
                            new_cell.value = '支持' if protect_type and protect_type != 'None' else '不支持'
                            if new_cell.value == '不支持':
                                new_cell.font = Font(color='FF0000')  # 设置红色字体
                            
                            if protect_ranges:
                                comment = Comment(protect_ranges, "System")
                                comment.width = 300
                                comment.height = 200
                                new_cell.comment = comment
                        elif col == first_rwe_column:
                            new_cell.value = '支持'
                        elif main_header in ['品牌', '主型号', 'MID', 'DID', '容量\n(MB)', '是否支持多倍速']:
                            new_cell.value = new_row.get(main_header, '')
                        elif main_header in ['料号', '封装', '最高时钟频率', '备注']:  # 从这里移除 'UID'
                            new_cell.value = ''  # 这些列保持空白
                        elif sub_header == 'UID':  # 专门处理 UID 子标题列
                            new_cell.value = ''  # UID 列保持空白
                        else:  # 其他有效列（列标题不为空）填入"不支持"并设置红色
                            new_cell.value = '不支持'
                            new_cell.font = Font(color='FF0000')  # 设置红色字体
                    
                    last_row = current_row
                
                del brand_groups[brand]
        
        # 添加剩余的新品牌到末尾
        for brand in brand_groups:
            for new_row in brand_groups[brand]:
                current_row = ws.max_row + 1
                ws.insert_rows(current_row)
                
                for col in range(1, ws.max_column + 1):
                    new_cell = ws.cell(row=current_row, column=col)
                    template_cell = ws.cell(row=data_start_row, column=col)
                    
                    if template_cell.has_style:
                        new_cell._style = copy(template_cell._style)
                    
                    if col == first_otp_column:
                        new_cell.value = new_row['OTP']
                    elif col == first_wp_column:
                        protect_type = new_row.get('protect_type', '')
                        protect_ranges = new_row.get('protect_ranges', '')
                        new_cell.value = '支持' if protect_type and protect_type != 'None' else '不支持'
                        
                        if protect_ranges:
                            comment = Comment(protect_ranges, "System")
                            comment.width = 300
                            comment.height = 200
                            new_cell.comment = comment
                    elif col == first_rwe_column:
                        new_cell.value = '支持'
                    else:
                        col_name = ws.cell(row=header_row, column=col).value
                        if not col_name:
                            col_name = ws.cell(row=subheader_row, column=col).value
                        if col_name in new_row and col_name != 'OTP':
                            new_cell.value = new_row[col_name]
                        else:
                            new_cell.value = ''
        
        # 保存修改到新文件
        wb.save(output_excel)
        print(f"成功添加了 {len(new_rows)} 个新的Flash设备")
        
    except Exception as e:
        print(f"更新Excel时发生错误：{str(e)}")
        import traceback
        print(traceback.format_exc())

def main():
    try:
        # 读取spinor_ids.c文件
        with open('drivers/spinor/spinor_ids.c', 'r') as f:
            content = f.read()
        
        # 解flash信息
        flash_list = parse_flash_info(content)
        
        # 比较并更新Excel文件
        compare_and_update_excel(
            flash_list, 
            'SPI_NOR_FLASH_V19.xlsx',
            'SPI_NOR_FLASH_V20.xlsx'
        )
    except FileNotFoundError:
        print("错误：找不到必要的文件。请确以下文件存在：")
        print("1. drivers/spinor/spinor_ids.c")
        print("2. SPI_NOR_FLASH_V19.xlsx")
    except Exception as e:
        print(f"发生错误：{str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == '__main__':
    # 首先检查并安装必要包
    try:
        import pkg_resources
        pkg_resources.require(['openpyxl'])
    except (pkg_resources.DistributionNotFound, pkg_resources.VersionConflict):
        print("正在安装要的Python包...")
        import subprocess
        subprocess.check_call(['pip3', 'install', 'openpyxl'])
        print("安装完成")
    
    main()

