import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
from copy import copy
import shutil
import argparse
import os

def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='生成 SPI NOR Flash 支持列表')
    parser.add_argument('-i', '--input', required=True,
                      help='输入的Excel文件路径')
    parser.add_argument('-o', '--output', required=True,
                      help='输出的Excel文件路径')
    return parser.parse_args()

class FlashParser:
    def __init__(self):
        pass

    def parse_flash_info(self, file_content):
        """解析所有flash信息并返回列表"""
        flash_list = []
        jedec_id_map = {}  # 用于检测重复的 JEDEC ID 和 Name

        # Flash信息解析的正则表达式
        pattern = r'\{\s*"([^"]+)",\s*'  # Flash名称
        pattern += r'0x([0-9a-f]+),\s*'   # JEDEC ID
        pattern += r'0x([0-9a-f]+),\s*'   # 容量
        pattern += r'([^,]+),\s*'         # 特性
        pattern += r'([^,]+),\s*'         # get_flash_markid
        pattern += r'([^,]+),\s*'         # markid
        pattern += r'([^,]+),\s*'         # protect
        pattern += r'([^}]+)\s*\},'       # otp

        matches = re.finditer(pattern, file_content, re.MULTILINE | re.DOTALL)

        for match in matches:
            flash_name = match.group(1)

            # 跳过不支持的Flash类型
            if flash_name == "Unsupported Flash Type":
                continue

            jedec_id_str = match.group(2)

            # 检查是否存在完全相同的设备(JEDEC ID和Name都相同)
            if jedec_id_str in jedec_id_map and flash_name in jedec_id_map[jedec_id_str]:
                print(f"\n警告: 发现重复的Flash设备:")
                print(f"  JEDEC ID: 0x{jedec_id_str}")
                print(f"  型号: {flash_name}")
                print("  跳过添加重复项...")
                print("")
                continue

            # 如果是相同JEDEC ID但不同Name的设备，则允许添加
            if not jedec_id_str in jedec_id_map:
                jedec_id_map[jedec_id_str] = set()

            # 记录该JEDEC ID下的型号
            jedec_id_map[jedec_id_str].add(flash_name)

            size = match.group(3)
            features = match.group(4).strip()
            get_markid = match.group(5).strip()
            markid = match.group(6).strip()
            protect = match.group(7).strip()
            otp = match.group(8).strip() if match.group(8) else ''

            flash_info = self._build_flash_info(
                flash_name, jedec_id_str, size, features,
                get_markid, markid, protect, otp, file_content
            )

            flash_list.append(flash_info)

        return flash_list

    def _parse_brand(self, flash_name, jedec_id):
        """解析品牌信息"""
        # 品牌映射字典
        brand_map = {
            'EN': 'Eon(宜杨)',
            'F': 'ESMT(晶豪科技)',
            'GD': 'GigaDevice(兆易创新)',
            'M': 'Micron(镁光)',
            'N': 'Micron(镁光)',
            'MX': 'MXIC(旺宏)',
            'KH': 'MXIC(旺宏)',
            'W': 'Winbond(华邦)',
            'S': 'Cypress(赛普拉斯)',
            'PM': 'PMC(常亿)',
            'PN': 'Paragon(芯天下曾用名)',
            'IC': 'ISSI(芯成)',
            'ZB': 'Zbit(恒烁)',
            'BY': 'Boya(博雅)',
            'BH': 'Boya(博雅)',
            'XM': 'XMC(武汉新芯)',
            'FM': 'FUDAN(复旦微)',
            'FS': 'FORESEE(江波龙)',
            'XT': 'XTX(芯天下)',
            'ZD': 'Zetta(芯泽)',
            'NM': 'NORMEM(诺存)',
            'DS': 'Doslicon(东芯)',
            'JY': 'Juyang(巨扬)',
            'UC': 'Ucun(优存科技)',
            'PY': 'Puya(普冉)',
            'T': 'First-Rank(飞思瑞克)',
            'SX': 'SANXIA(三峡)',
            'GT': 'GT(聚辰)',
            'TH': 'TH(紫光青藤)',
            'BW': 'BW(佰维)'
        }

        # 特殊情况处理
        if flash_name == 'FM25Q64A' and jedec_id.lower() == '0x00f83217':
            return 'Doslicon(东芯)'

        prefix = flash_name[:2] if len(flash_name) >= 2 else flash_name[:1]

        if prefix in brand_map:
            return brand_map[prefix]

        if len(prefix) == 2 and prefix[0] in brand_map:
            return brand_map[prefix[0]]

        # 找不到品牌映射时提示用户
        print(f"\n警告: 未找到Flash型号 {flash_name} 对应的品牌映射")
        print(f"请在 brand_map 中添加相应的品牌映射")

        return flash_name.split()[0]

    def _parse_features(self, features):
        """解析特性信息"""
        feature_list = []
        if 'SPI_NOR_DUAL_READ' in features:
            feature_list.append('DUAL_READ')
        if 'SPI_NOR_QUAD_READ' in features:
            feature_list.append('QUAD_READ')
        if 'SPI_NOR_QUAD_WRITE' in features:
            feature_list.append('QUAD_WRITE')
        return '/'.join(feature_list) if feature_list else 'None'

    def _parse_protect_info(self, protect, file_content):
        """解析写保护范围信息"""
        # 如果protect为空或NULL,直接返回空字符串
        if not protect or protect == 'NULL':
            return ''

        # 解析保护类型
        protect_match = re.search(r'&protect_([a-z0-9_]+)', protect)
        if not protect_match:
            return ''

        protect_type = protect_match.group(1)

        # 解析写保护范围
        pattern = rf'static\s+struct\s+spi_nor_flash_protect_range\s+protect_range_{protect_type}\[\]\s*=\s*\{{([^;]+)\}};'
        protect_def_match = re.search(pattern, file_content, re.DOTALL)
        if not protect_def_match:
            return ''

        array_content = protect_def_match.group(1)
        ranges = []

        entries = re.finditer(r'\{\s*0x[0-9a-fA-F]+,\s*0x[0-9a-fA-F]+,\s*0x[0-9a-fA-F]+,\s*0x[0-9a-fA-F]+,\s*(\d+)\s*\*\s*0x10000\s*\}', array_content)

        for entry in entries:
            try:
                blocks = int(entry.group(1))
                if blocks == 0:
                    end_addr = 0
                else:
                    end_addr = blocks * 0x10000 - 1

                range_str = f"{blocks:2d} blocks   : 0x00 ~ 0x{end_addr:06X}"
                ranges.append(range_str)
            except (ValueError, IndexError):
                continue

        # 按块数排序
        ranges.sort(key=lambda x: int(x.split()[0]))
        return '\n'.join(ranges)

    def _parse_otp_info(self, otp_str, file_content, flash_name):
        """解析OTP信息"""
        if not otp_str or 'NULL' in otp_str:
            return ''

        otp_type_match = re.search(r'otp_type(\d+)', otp_str)
        if not otp_type_match:
            return ''

        otp_type_num = otp_type_match.group(1)

        pattern = rf'static\s+struct\s+spi_nor_otp\s+otp_type{otp_type_num}\s*=\s*\{{([^}}]+)\}}'
        otp_def_match = re.search(pattern, file_content)
        if not otp_def_match:
            return ''

        otp_values = otp_def_match.group(1).split(',')
        if len(otp_values) >= 4:
            try:
                regions = int(otp_values[-2].strip())
                size = int(otp_values[-3].strip())
                return f"{regions}*{size} Byte"
            except ValueError:
                return ''

        return ''

    def _build_flash_info(self, flash_name, jedec_id_str, size, features,
                         get_markid, markid, protect, otp, file_content):
        """构建单个Flash信息"""
        # 解析各项信息
        jedec_id = int(jedec_id_str, 16)
        size_mb = int(size, 16) / (1024 * 1024)
        features_str = self._parse_features(features)
        protect_ranges = self._parse_protect_info(protect, file_content)
        otp_desc = self._parse_otp_info(otp, file_content, flash_name)
        brand = self._parse_brand(flash_name, f'0x{jedec_id_str}')

        return {
            'Flash Name': flash_name,
            'JEDEC ID': f'0x{jedec_id_str}',
            'Size(MB)': size_mb,
            'Features': features_str,
            'Get MarkID': 'Yes' if get_markid != 'NULL' else 'No',
            'Mark ID': markid if markid != '0' else 'None',
            'OTP': otp_desc,
            'Brand': brand,
            'Protect Ranges': protect_ranges
        }

class ExcelHandler:
    def analyze_excel_structure(self, excel_file):
        """解析Excel表格结构"""
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # 获取表头信息
            header_row = 1
            subheader_row = 2
            columns = {}

            print("\n=== Excel表格结构分析 ===")
            print("列号\t主标题\t子标题")
            print("-" * 40)

            for col in range(1, ws.max_column + 1):
                main_header = ws.cell(row=header_row, column=col).value
                sub_header = ws.cell(row=subheader_row, column=col).value

                if main_header or sub_header:
                    columns[col] = {
                        'main_header': main_header,
                        'sub_header': sub_header
                    }
                    print(f"{col}\t{main_header}\t{sub_header}")

            return columns

        except Exception as e:
            print(f"解析Excel结构时发生错误: {str(e)}")
            return None

    def compare_and_update_excel(self, flash_list, existing_excel, output_excel):
        """比较并更新Excel内容"""
        try:
            # 复制文件
            shutil.copy2(existing_excel, output_excel)

            # 加载工作簿并升级版本
            wb = openpyxl.load_workbook(output_excel)
            if not self.upgrade_version(wb):
                print("版本升级失败")
                return
            wb.save(output_excel)

            # 处理新的flash设备
            wb = openpyxl.load_workbook(output_excel)
            ws = wb.active

            # 定义头
            header_row = 1
            subheader_row = 2
            data_start_row = 3

            # 找到第一个OTP写保护和读写擦列
            first_otp_column = None
            first_wp_column = None
            first_rwe_column = None

            for col in range(1, ws.max_column + 1):
                cell_value = self.get_merged_cell_value(ws, ws.cell(row=subheader_row, column=col))
                main_header = self.get_merged_cell_value(ws, ws.cell(row=header_row, column=col))

                if cell_value:
                    cell_value = str(cell_value).strip()
                    if cell_value == 'OTP' and not first_otp_column:
                        first_otp_column = col
                    elif cell_value == '写保护' and not first_wp_column:
                        first_wp_column = col
                    elif cell_value == '读写擦' and not first_rwe_column:
                        first_rwe_column = col

            if not first_otp_column or not first_wp_column or not first_rwe_column:
                print("错误: 未找到必要的列")
                print(f"OTP列: {first_otp_column}")
                print(f"写保护列: {first_wp_column}")
                print(f"读写擦列: {first_rwe_column}")
                return

            # 获取现有的MID-DID组合
            existing_mid_did = set()
            for row in range(data_start_row, ws.max_row + 1):
                mid = str(self.get_merged_cell_value(ws, ws.cell(row=row, column=5))).lower()
                did = str(self.get_merged_cell_value(ws, ws.cell(row=row, column=6))).lower()
                if mid != "none" and did != "none":
                    existing_mid_did.add(f"{mid}-{did}")

            # 处理新的Flash设备
            new_rows = []
            inserted_devices = []  # 用于记录新增的设备
            for flash in flash_list:
                jedec_id = flash['JEDEC ID'].lower()
                if len(jedec_id) >= 10:
                    jedec_id = "0x" + jedec_id[4:]

                mid = f"0x{jedec_id[2:4].upper()}"
                did = f"0x{jedec_id[4:8].upper()}"

                mid_did = f"{mid.lower()}-{did.lower()}"
                if mid_did in existing_mid_did:
                    continue

                brand = flash['Brand']

                # 定义speed_support_str
                features = str(flash['Features'])
                if ('DUAL_READ' in features and
                    'QUAD_READ' in features and
                    'QUAD_WRITE' in features):
                    speed_support_str = '支持'
                elif 'DUAL_READ' in features:
                    speed_support_str = '不支持4倍速'
                else:
                    speed_support_str = '不支持2、4倍速'

                new_rows.append({
                    '品牌': brand,
                    '主型号': flash['Flash Name'],
                    '料号': '',
                    '封装': 'SOP',  # 默认封装为 SOP
                    'MID': mid,
                    'DID': did,
                    '容量\n(MB)': int(flash['Size(MB)']),
                    '是否支持多倍速': speed_support_str,
                    '最高时钟频率': '',
                    '备注': '',
                    'OTP': flash['OTP'] if 'OTP' in flash else '',
                    'protect_ranges': flash['Protect Ranges']  # 只保留写保护范围信息
                })

            # 按品牌分组
            brand_groups = {}
            for new_row in new_rows:
                brand = new_row['品牌']
                if brand not in brand_groups:
                    brand_groups[brand] = []
                brand_groups[brand].append(new_row)

            # 相应品牌后插入新行
            current_row = data_start_row
            for row in range(data_start_row, ws.max_row + 1):
                brand = self.get_merged_cell_value(ws, ws.cell(row=row, column=1))
                if brand in brand_groups:
                    # 找到品牌的所有行
                    brand_rows = []
                    last_row = row
                    while (last_row < ws.max_row and
                           self.get_merged_cell_value(ws, ws.cell(row=last_row + 1, column=1)) == brand):
                        last_row += 1

                    # 获取该品牌所有行的容量
                    brand_capacities = []
                    for r in range(row, last_row + 1):
                        capacity = self.get_merged_cell_value(ws, ws.cell(row=r, column=7))  # 容量列
                        if capacity:
                            brand_capacities.append((r, float(capacity)))

                    # 对该品牌的新设备按容量排序
                    new_devices = brand_groups[brand]
                    new_devices.sort(key=lambda x: float(x['容量\n(MB)']))

                    # 根据容量大小插入新设备
                    offset = 0  # 用于记录已插入的行数
                    for new_row in new_devices:
                        new_capacity = float(new_row['容量\n(MB)'])

                        # 找到合适的插入位置
                        insert_pos = last_row + 1  # 默认插在最后
                        next_bigger_capacity_pos = None  # 记录下一个更大容量的位置

                        for row_idx, capacity in brand_capacities:
                            if capacity > new_capacity:
                                if next_bigger_capacity_pos is None:
                                    next_bigger_capacity_pos = row_idx
                                break

                        # 如果找到更大容量的位置，插入到它前面
                        if next_bigger_capacity_pos is not None:
                            # 如果之前插入的行在当前要插入的位置之前，需要向后偏移
                            if next_bigger_capacity_pos > row:
                                insert_pos = next_bigger_capacity_pos + offset
                            else:
                                insert_pos = next_bigger_capacity_pos
                        else:
                            insert_pos = last_row + 1 + offset


                        # 更新offset
                        offset += 1

                        # 插入新行并复制样式
                        ws.insert_rows(insert_pos)

                        # 每个新的 flash 设备重置 UID 标记
                        is_first_uid = True  # 在这里初始化变量

                        # 复制上一行的样式
                        for col in range(1, ws.max_column + 1):
                            new_cell = ws.cell(row=insert_pos, column=col)
                            prev_cell = ws.cell(row=insert_pos - 1, column=col)  # 获取上一行的单元格

                            # 复制上一行的样式
                            if prev_cell.has_style:
                                new_cell._style = copy(prev_cell._style)
                            if prev_cell.font:
                                new_cell.font = copy(prev_cell.font)
                            if prev_cell.border:
                                new_cell.border = copy(prev_cell.border)
                            if prev_cell.fill:
                                new_cell.fill = copy(prev_cell.fill)
                            if prev_cell.alignment:
                                new_cell.alignment = copy(prev_cell.alignment)

                            # 获取标题
                            main_header = self.get_merged_cell_value(ws, ws.cell(row=header_row, column=col))
                            sub_header = self.get_merged_cell_value(ws, ws.cell(row=subheader_row, column=col))

                            # 如果标题和子标题为空，跳过该列
                            if not main_header and not sub_header:
                                continue

                            # 设置值
                            if col == first_otp_column:
                                if new_row['OTP']:
                                    new_cell.value = new_row['OTP']
                                else:
                                    new_cell.value = '不支持'
                                    new_cell.font = Font(color='FF0000')  # 设置红色字体
                            elif col == first_wp_column:
                                protect_ranges = new_row.get('protect_ranges', '')
                                new_cell.value = '支持' if protect_ranges and protect_ranges != 'None' else '不支持'
                                if new_cell.value == '不支持':
                                    new_cell.font = Font(color='FF0000')  # 设置红色字体

                                if protect_ranges:
                                    comment = Comment(protect_ranges, "System")
                                    comment.width = 300
                                    comment.height = 200
                                    new_cell.comment = comment
                            elif col == first_rwe_column:
                                new_cell.value = '支持'
                            elif main_header == '容量\n(MB)':  # 修改这里，专门处理容量列
                                new_cell.value = new_row['容量\n(MB)']
                            elif main_header in ['品牌', '主型号', 'MID', 'DID', '是否支持多倍速']:  # 从这里移除容量列
                                new_cell.value = new_row.get(main_header, '')
                            elif main_header in ['料号', '最高时钟频率', '备注']:  # 从这里移除封装，因为封装有默认值
                                new_cell.value = ''  # 这些列保持空白
                            elif main_header == '封装':  # 专门处理封装列
                                new_cell.value = 'SOP'  # 设置默认封装为 SOP
                            elif sub_header == 'UID':  # 专门处理 UID 子标题列
                                if is_first_uid:  # 第一个 UID 列保持空白
                                    new_cell.value = ''
                                    is_first_uid = False  # 标记已处理第一个 UID 列
                                else:  # 其他 UID 列填入红色的"不支持"
                                    new_cell.value = '不支持'
                                    new_cell.font = Font(color='FF0000')  # 设置红色字体
                            else:  # 其他有效列（列标题不为空）填入"不支持"并设置红色
                                new_cell.value = '不支持'
                                new_cell.font = Font(color='FF0000')  # 设置红色字体

                        # 记录新增设备信息
                        capacity = new_row['容量\n(MB)']
                        inserted_devices.append({
                            'name': new_row['主型号'],
                            'jedec_id': f"0x{new_row['MID'][2:]}{new_row['DID'][2:]}",
                            'size': f"{capacity}MB",
                            'row': insert_pos
                        })

                        insert_pos += 1

                    # 从已处理的品牌组中移除
                    del brand_groups[brand]

            # 处理剩余的未匹配品牌
            if brand_groups:
                # 在文件末尾添加剩余的设备
                insert_pos = ws.max_row + 1
                for brand, devices in brand_groups.items():
                    for new_row in devices:
                        # ... (与上面相同的插入逻辑)
                        pass

            # 保存修改
            wb.save(output_excel)

            # 打印新增设备信息
            if inserted_devices:  # 只有在有新增设备时才打印
                print("\n=== 新增的Flash设备 ===")
                print("-" * 70)
                print(f"{'型号':<22} {'JEDEC ID':<13} {'容量':<6} {'Excel 行号':>12}")
                print("-" * 70)
                for device in inserted_devices:
                    print(f"{device['name']:<24} {device['jedec_id']:<13} {device['size']:<6} {device['row']:>12}")
                print("-" * 70)

                print("\n请注意：")
                print("1. 所有新增设备默认封装为 SOP，如果实际封装不是 SOP，请在 Excel 中手动修改")
                print("2. 需要为新增设备补充 料号、最高时钟频率、UID")

        except FileNotFoundError as e:
            print(f"\n错误: 找不到文件 - {str(e)}")
        except PermissionError as e:
            print(f"\n错误: 没有足够的权限 - {str(e)}")
        except Exception as e:
            print(f"\n错误: {str(e)}")
        finally:
            try:
                wb.save(output_excel)
            except Exception:
                pass

    def upgrade_version(self, wb):
        """升级Excel版本号"""
        ws = wb.active
        header_row = 1
        subheader_row = 2

        # 找到备注列
        remark_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col).value == '备注':
                remark_col = col
                break

        if not remark_col:
            print("未找到备注列")
            return False

        src_start_col = remark_col + 1

        # 保存所有合并单元格信息
        merged_ranges = []
        for merged_range in ws.merged_cells.ranges:
            merged_ranges.append({
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col,
                'value': ws.cell(merged_range.min_row, merged_range.min_col).value
            })

        # 保存所有列的内容
        all_columns_data = []
        for col in range(src_start_col, ws.max_column + 1):
            # 检查列是否有效（标题行或子标题行有内容）
            if not (ws.cell(row=header_row, column=col).value or
                    ws.cell(row=subheader_row, column=col).value):
                break

            col_data = []
            for row in range(1, ws.max_row + 1):
                src_cell = ws.cell(row=row, column=col)
                cell_info = {
                    'value': src_cell.value,
                    'style': copy(src_cell._style) if src_cell.has_style else None,
                    'font': copy(src_cell.font) if src_cell.font else None,
                    'border': copy(src_cell.border) if src_cell.border else None,
                    'fill': copy(src_cell.fill) if src_cell.fill else None,
                    'alignment': copy(src_cell.alignment) if src_cell.alignment else None,
                    'comment': copy(src_cell.comment) if src_cell.comment else None
                }
                col_data.append(cell_info)
            all_columns_data.append(col_data)

        # 插入新版本的4列
        ws.insert_cols(src_start_col, 4)

        # 复制最高版本的4列内容到新位置
        temp_data = all_columns_data[:4]
        for col_idx, col_data in enumerate(temp_data):
            dst_col = src_start_col + col_idx

            for row_idx, cell_info in enumerate(col_data, 1):
                dst_cell = ws.cell(row=row_idx, column=dst_col)

                # 设置值（只修改版本号）
                if row_idx == header_row and col_idx == 0:
                    version_str = cell_info['value']
                    if version_str and isinstance(version_str, str) and version_str.startswith('V'):
                        version = version_str.replace('V', '')
                        major, minor, patch = map(int, version.split('.'))
                        new_version = f"V{major}.{minor+1}.0"
                        dst_cell.value = new_version
                        print(f"版本升级: {version_str} -> {new_version}")
                else:
                    dst_cell.value = cell_info['value']

                # 恢复样式和批注
                for attr in ['style', 'font', 'border', 'fill', 'alignment']:
                    if cell_info[attr]:
                        setattr(dst_cell, attr if attr != 'style' else '_style', cell_info[attr])

                # 恢复批注
                if cell_info['comment']:
                    dst_cell.comment = cell_info['comment']

        # 恢复所有合并单元格
        ws.merged_cells.ranges.clear()
        for merged_range in merged_ranges:
            min_col = merged_range['min_col']
            max_col = merged_range['max_col']

            if min_col >= src_start_col:
                # 移动合并单元格
                new_min_col = min_col + 4
                new_max_col = max_col + 4
            else:
                # 保持原有合并单元格
                new_min_col = min_col
                new_max_col = max_col

            new_range = f"{get_column_letter(new_min_col)}{merged_range['min_row']}:{get_column_letter(new_max_col)}{merged_range['max_row']}"
            ws.merge_cells(new_range)

        # 合并新版本号单元格
        merge_range = f"{get_column_letter(src_start_col)}{header_row}:{get_column_letter(src_start_col+3)}{header_row}"
        ws.merge_cells(merge_range)

        return True

    def get_merged_cell_value(self, ws, cell):
        """获取合并单元格的值"""
        if isinstance(cell, MergedCell):
            # 找到这个合并单元格所属的范围
            for merged_range in ws.merged_cells.ranges:
                if (cell.row >= merged_range.min_row and cell.row <= merged_range.max_row and
                    cell.column >= merged_range.min_col and cell.column <= merged_range.max_col):
                    # 返回合并单元格左上角的值
                    return ws.cell(merged_range.min_row, merged_range.min_col).value
        return cell.value

def main():
    # 解析命令行参数
    args = parse_arguments()

    # 使用固定的相对路径
    spinor_ids_path = '../../drivers/spinor/spinor_ids.c'

    try:
        # 检查文件是否存在
        if not os.path.exists(spinor_ids_path):
            print(f"\n错误: 找不到源文件 {spinor_ids_path}")
            return

        if not os.path.exists(args.input):
            print(f"\n错误: 找不到输入文件 {args.input}")
            return

        # 检查输出文件是否已存在
        if os.path.exists(args.output):
            response = input(f"\n警告: 输出文件 {args.output} 已存在。是否覆盖？(y/N): ")
            if response.lower() != 'y':
                print("操作已取消")
                return

        # 读取并解析flash信息
        with open(spinor_ids_path, 'r', encoding='utf-8') as f:
            content = f.read()

        flash_parser = FlashParser()
        flash_list = flash_parser.parse_flash_info(content)

        # 更新Excel文件
        excel_handler = ExcelHandler()
        excel_handler.compare_and_update_excel(
            flash_list,
            args.input,
            args.output
        )

    except Exception as e:
        print(f"\n错误: {str(e)}")

if __name__ == '__main__':
    main()

